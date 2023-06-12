#!/bin/python
import os
import sys
import argparse


    

'''# 查询issue信息，传入参数issueId
issue = jiraClinet.issue('xxx-679')

# 问题的 key
issue.key

# 问题的 id
issue.id

# 问题的配置域
issue.fields

# 问题标题描述
issue.fields.summary

# 问题详细描述
issue.fields.description

# 问题的类型
issue.fields.issuetype

#问题报告者
issue.fields.reporter '''

import sys

import argparse
  
import requests
import re
import openpyxl
import datetime
from openpyxl.worksheet.worksheet import Worksheet
from basic_auth import *

def getCellValue(src, row, col):
    return src.cell_value(row, col)

# def XlsCharToInt(col):
#     if type(col) == str:
#         return ord(col) - ord('A')
#     return col

def getJiraShellCount(wbook,shellName):
    wshell = wbook[shellName]
    assert isinstance(wshell,Worksheet)
    jiraCount=[]
    
    # for row in range(wshell.max_row):
    for col in range(wshell.max_column):
        try:
            if col == 0:
                jiraCount.append(datetime.datetime.now().strftime("%Y/%m/%d"))
                continue

            link = wshell.cell(4,col+1)
            jiraCount.append(0)
            if link.hyperlink == None: continue
            url = link.hyperlink.target #使用这个可以，使用xlrd不行
            response = requests.get(url,auth=basic_auth)
            content = response.content
            c_s_i_e = r'<span class="results-count-total results-count-link">.*</span>'
            contents = re.findall(c_s_i_e,str(content),re.A)
            e_i=r"-?\b[a-zA-Z_0x0-9.]+\b" 
            if len(contents) !=0:
                contents = re.findall(e_i,str(contents[0]),re.A)
                print(link.value,contents[8])
                jiraCount[col] = int(contents[8])
        except:
            print(f"{col}行解析错误")
            pass

    isWrite = False
    for row in range(wshell.max_row):
        try:
            if row >= 5 :
                cell = wshell.cell(row+1,1)
                if str(cell.value) == "" or cell.value == None:
                    for count in range(len(jiraCount)):
                        wshell[row][count].value = jiraCount[count]
                        isWrite = True
                    print("处理完1"+shellName)
            if isWrite:
                break
        except:
            continue

    if not isWrite:
        wshell.append(jiraCount)
        print("处理完2"+shellName)
   
    
def getJiraCount(xlsFileName):
    wbook = openpyxl.load_workbook(xlsFileName)
    getJiraShellCount(wbook,"Bug")
    getJiraShellCount(wbook,"Task")
    getJiraShellCount(wbook,"客户反馈")
    wbook.save(xlsFileName)
    print("生成完成")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
    description='''
    获取Jira上的信息
    ''')
    
    #这个是要解析 -f 后面的参数
    parser.add_argument('-f', '--xlsFileName',help="文件名称",type=str,default='',nargs='?')
    arg=parser.parse_args()

    getJiraCount(arg.xlsFileName)
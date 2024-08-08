#!/usr/bin/python
import argparse
from analyze_dbc.commonfun import*
from analyze_dbc.analyze_dbc import *
from analyze_dbc.projectInI import *
from xlrd.book import Book
from xlrd.sheet import Sheet
import xlrd
import openpyxl
import copy

'''
    可以检测的前提条件
    message名称必须是id结尾
'''

ISXLS = 'isXls'
LOGSTR = 'log'
def findsignalInfile(signal,content):
    try:
        for text in content:
            texts = re.findall(e_i,text,re.A)
            try:
                if signal == texts[0]:
                    return True
            except:
                pass
        return False
    except:
        return False

def getJsConfig(configPath="",down_config="",up_config="",isPrint = False):
    if configPath == "": configPath = pyFileDir+"config.json"
    jsConfig = getJScontent(configPath)
    dbcfile = getKeyPath("dbcfile", jsConfig)

    if down_config == "":
        down_config = getKeyPath("down", jsConfig)
    if up_config == "":
        up_config = getKeyPath('up', jsConfig)
    
    if isPrint:
        print("down_config:",down_config)
        print("up_config:",up_config)
    print(dbcfile)
    dbc = Analyze(dbcfile)
    jsDown = getJScontent(down_config)
    jsUp = getJScontent(up_config)
    return jsDown,jsUp,dbc,jsConfig,down_config,up_config

def CheckSigName(configPath,down_config="",up_config=""):
    jsDown, jsUp, dbc, jsConfig ,down_config,up_config= getJsConfig(configPath,down_config,up_config,True)
    sigNames = {}

    igSigs = []
    checkSigIgnore = getKeyPath("checkSigIgnore",jsConfig)
    if os.path.exists(checkSigIgnore):
        igSigs = getJScontent(checkSigIgnore)
    for top in jsDown:
        topValue = jsDown[top]
        topNoSet = EesyStr.removeAll(top, SETSTR)
        if type(topValue) == str:
            if topValue != top:
                sigNames[topValue] = 0
        else:
            for t in topValue:
                if t != top and t != topNoSet:
                    sigNames[t] = 0
    for s in jsUp:
        isTopc = False
        if type(jsUp[s]) == dict:
            for k in jsUp[s]:
                if k == "bindSigNames":
                    isTopc = True
                    for bindSigName in jsUp[s][k]:
                        sigNames[bindSigName] = s
        if not isTopc: sigNames[s] = 1

    for igSig in igSigs:
        if igSig in sigNames:
            del sigNames[igSig]

    bar= ProgressBar()
    bar.total = len(sigNames)
    igCmd =""
    for sig in sigNames:
        bar.printCurrnt()
        if is_chinese(sig):
            continue
        dbcSigName,Sender,isChanged,messageId = configConverdbc(sig,dbc)

        if Sender == None:
            if 'd' not in igCmd:
                igCmd = input("下次检测是否忽略y(是)/n(否)/dy(全是)/dn(全否)")
            if 'y' in igCmd:
                igSigs.append(sig)
            continue
        if  not isChanged:
            continue
        can_parse_whitelistPath = getKeyPath("can_parse_whitelist", jsConfig)
        f=open(can_parse_whitelistPath,'r')
        whitelistPath_content=f.readlines()
        f.close()
        messagesig = dbcSigName.replace("/","__")
        if  not findsignalInfile(f'{messagesig}',whitelistPath_content):
            printRed(f'{messagesig} 不在白名单中')
        
        if  sigNames[sig] == 1: 
            addConfigDict(jsUp,dbcSigName,sig)       
        elif  sigNames[sig] == 0: 
            addConfigDict(jsDown,sig,dbcSigName)
        elif type(sigNames[sig]) == str:
            replaceIndex = jsUp[sigNames[sig]]["bindSigNames"].index(sig)
            jsUp[sigNames[sig]]["bindSigNames"][replaceIndex] = dbcSigName

    writeJs(down_config,jsDown)
    writeJs(up_config,jsUp)
    writeJs(checkSigIgnore,igSigs)
    bar.printCurrnt()
    printGreen("分析完成")

def ToJaveCode(javaPath,configPath,down_config=None,up_config=None):
    pass


def splitSig(sig):
    assert isinstance(sig,str)
    sigs = []
    sigs.append(sig)
    if "/" in sig:
        tempSigs =  sig.split("/")
    elif "__" in sig:
        tempSigs =  sig.split("__")
    else:
        return sigs

    if len(tempSigs) < 2:
        return sigs
    sigs.clear()
    tempMesages = re.findall(m_s,tempSigs[0])
    id = tempMesages.pop()
    sender = "_".join(tempMesages)
    sigs.append(sender)
    sigs.append(id)
    sigs.append(tempSigs[1])
    return sigs
        

def set_highest_bit_to_one(num):
    # 确保 num 是一个 32 位整数
    assert isinstance(num, int)

    if num > 0x7FF:
        # 使用按位或操作符将最高位置为 1
        num |= 0x80000000
    return num

def configConverdbc(sig,dbc,isPrint=True):
    assert isinstance(dbc,Analyze)
    assert isinstance(sig,str)
    sigs = splitSig(sig)
    dbcSig = None
    messageReplace = True
    try:
        messageReplace = MESSAGEREPLACE
    except:
        pass
    if len(sigs) >= 3:
        try:
            sig16= str(set_highest_bit_to_one(int(sigs[1],16)))+"_".join(sigs[2:])
            prefix= f"{sigs[0]}_{sigs[1]}"
            dbcSig = dbc.getSig(sig16,sigs[0])
            if dbcSig != None and dbcSig.getMessage_Name() != prefix:
                if messageReplace:
                    if isPrint:
                        printYellow(f' {sig:<50} {prefix} {dbcSig.getMessage_Name()} message名称错误，即将替换')
                    return sig.replace(prefix,dbcSig.getMessage_Name()),dbcSig.Sender,True,dbcSig.messageId
                else:
                    dbcSig=None
        except:
            pass
    if dbcSig==None:
        dbcSig = dbc.getSig(sig)
        if dbcSig == None and isPrint:
            printRed(f'{sig:<50}  信号格式错误')
    return sig if dbcSig == None else dbcSig.getMessage_Name()+'/'+dbcSig.name, None if dbcSig == None else dbcSig.Sender,False,None if dbcSig == None else dbcSig.messageId

def addConfigDict(js,key,value):
    if value in js: del js[value] 
    js[key] = value

def reConfigDict(js,key,value): # 添加进容器，如果有重复就打印，不添加
    if key in js :
        try:
            if js[key] == value:
                printYellow(f'{value} 是存在的')
                return
            else:
                js[key] = value
        except:
            js[key] = value
    else:
        js[key] = value
    printGreen(f'{value} 添加完成')

def addSet(topic):
    if not topic.endswith(SETSTR): topic = topic+SETSTR
    return topic

def ToConfigJson(javaPath,configPath,down_config="",up_config=""):
    contents = readFileLines(javaPath)
    jsDown, jsUp, dbc, jsConfig,down_config,up_config = getJsConfig(configPath,down_config,up_config)
    for content in contents:
        tContents = re.findall(r'topic = ".*"',content)
        if len(tContents) != 0:
            tContent = re.findall(r'".*"',tContents[0])[0]
            assert isinstance(tContent,str)
            tContent=tContent.replace('\"',"")
            dbcSigName,Sender,isChanged,messageId = configConverdbc(tContent,dbc)
            down_topic=addSet(tContent)
            if Sender == None:
                continue
            if Sender not in local_machine_Sender: 
                addConfigDict(jsUp,dbcSigName,tContent)       
                # if down_topic in jsDown :del jsDown[down_topic]   
            if Sender in local_machine_Sender: 
                addConfigDict(jsDown,down_topic,dbcSigName)
                # if dbcSigName in jsUp: del jsUp[dbcSigName]

    writeJs(down_config,jsDown)
    writeJs(up_config,jsUp)

def addConfigSig(sigs,isOriginal,configPath="",down_config="",up_config=""):
    jsDown, jsUp, dbc, jsConfig,down_config,up_config = getJsConfig(configPath,down_config,up_config)
    for sigName in sigs:
        assert isinstance(sigName,str)
        dbcSigName,Sender,isChanged,messageId = configConverdbc(sigName,dbc)
        if isOriginal or SETSTR in sigName:
            down_topic = addSet(sigName)
        else:
            down_topic = addSet(dbcSigName)
        
        if isOriginal or SETSTR in sigName:
            up_topic = sigName
        else:
            up_topic = dbcSigName
            
        if Sender == None:
            continue
        if Sender not in local_machine_Sender: 
            reConfigDict(jsUp,dbcSigName,up_topic)       
            # if down_topic in jsDown :del jsDown[down_topic]   
        if Sender in local_machine_Sender: 
            reConfigDict(jsDown,down_topic,dbcSigName)
            # if dbcSigName in jsUp: del jsUp[dbcSigName]


    writeJs(down_config,jsDown)
    writeJs(up_config,jsUp)

def getCellValue(src, row, col):
    return src.cell_value(row,XlsCharToInt(col))

class SigXls():
    def __init__(self) -> None:
        self.sigName =""
        self.msgid = ""
        self.Sender = ""

    def msgId_sigName(self):
        return self.Sender+"_"+self.msgid+'/'+self.sigName

class ConfigXls():
    def __init__(self) -> None:
        self.topic = ""
        self.sigName = ""
        self.dbcSigName =""
        self.comments = ""
        self.valueMap = {}
        self.defaultValueType = ""
        self.sender = ""
        self.log = 1
        self.sigNameType = 0 #0是带msg,1只是信号名称

    def addToConfig(self,jsConfig,isDown,noChangedTopic):
        contentJson = {}
        if self.topic == '':
            self.topic = self.dbcSigName

        if not noChangedTopic:
            if isDown:
                if not self.topic.endswith(SETSTR): self.topic = self.topic +SETSTR
            else:
                if self.topic.endswith(SETSTR): self.topic = self.topic.rstrip(SETSTR)
        if len(self.valueMap) != 0:
            contentJson['valueMap'] = self.valueMap
        if len(self.defaultValueType) != 0:
            contentJson['defaultValueType'] = int(self.defaultValueType)
        if len(self.comments) != 0:
            contentJson['comments'] = self.comments
        contentJson[ISXLS] = True
        if isDown:
            configSigName = self.dbcSigName
            if self.sigNameType == 1:
                configSigName = self.sigName

            if self.topic in jsConfig:
                exTopic = jsConfig[self.topic]
                if type(exTopic) == str:
                    del jsConfig[self.topic]
                    jsConfig[self.topic]={exTopic:{}}

            if self.topic not in jsConfig:
                jsConfig[self.topic] = {}
            jsConfig[self.topic][configSigName] = contentJson
        else:
            if self.dbcSigName in jsConfig:
                exTopic = jsConfig[self.dbcSigName]
                if type(exTopic) == str:
                    del jsConfig[self.dbcSigName]
                    jsConfig[self.dbcSigName]={exTopic:{}}
            if self.dbcSigName not in jsConfig:
                jsConfig[self.dbcSigName] = {}
            jsConfig[self.dbcSigName][LOGSTR] = self.log
            jsConfig[self.dbcSigName][self.topic] = contentJson

    def getList(self):
        newSheel = []
        newSheel.append(self.comments)
        newSheel.append(self.topic)
        newSheel.append(self.dbcSigName)
        return newSheel

    def getBindSigNames(self,sigName,dbc,whitelistdbcSigNames,jsConfig):
        if sigName.startswith('0x'):
            assert isinstance(dbc,Analyze)
            msgId = getNoOx16(sigName)
            sigs = dbc.getSigsByMsgId(msgId)
            bindSigNames=[]
            for dbcSig in sigs:
                assert isinstance(dbcSig,SigInfo)
                dbcSigName = dbcSig.getMessage_Name()+'/'+dbcSig.name
                bindSigNames.append(dbcSigName)
                whitelistdbcSigNames.append(dbcSig.name)
            if self.topic not in jsConfig:
                jsConfig[self.topic] = {}
            jsConfig[self.topic][LOGSTR] = self.log
            jsConfig[self.topic]['bindSigNames'] = bindSigNames
            return True
        return False
            
def getSigXls(xlsFileName):
    sigs = []
    book = xlrd.open_workbook(xlsFileName)
    assert isinstance(book, Book)
    sheel = book.sheet_by_name(Sig_Matrix)
    for i in range(sheel.nrows):
        try:
            sigName = getCellValue(sheel,i,XlsCharToInt('C'))
            if len(sigName) == 0:
                printRed(f'{i}行是空的')
                continue
            msgid =  str( getCellValue(sheel,i,XlsCharToInt('E'))).split(".")[0].replace('0x', '')
            Sender= getCellValue(sheel,i,XlsCharToInt('B'))
            sigXls = SigXls()
            sigXls.Sender = Sender
            sigXls.msgid = msgid
            sigXls.sigName = sigName
            sigs.append(sigXls)
        except:
            printRed(f'获取 {sigName} 信息失败')
            pass
    return sigs

def addConfigByCanXls(xlsFileName):
    sigs = []
    sigXls = getSigXls(xlsFileName)
    for sig in sigXls:
        assert isinstance(sig,SigXls)
        sigs.append(sig.msgId_sigName())
    addConfigSig(sigs,True)

def addMultipleSig(canmatrix,msgIds,topics):
    jsDown, jsUp, dbc, jsConfig,down_config,up_config = getJsConfig()
    if canmatrix =='':
        canmatrix = getKeyPath("canmatrix", jsConfig)
    sigXls = getSigXls(canmatrix)
    sigs = []
    for msgIdIndex in range(len(msgIds)):
        msgId = msgIds[msgIdIndex]
        for sig in sigXls:
            assert isinstance(sig,SigXls)
            if sig.msgid == msgId:
                sigs.append(sig.msgId_sigName())

        if len(topics)!=0:
            bindSigNames=[]
            for sigName in sigs:
                assert isinstance(sigName,str)
                try:
                    dbcSigName,Sender,isChanged,messageId = configConverdbc(sigName,dbc)
                except:
                    printRed(f"{sigName} 信号格式错误")
                    continue   
                bindSigNames.append(dbcSigName)

            if Sender not in local_machine_Sender and not topics[msgIdIndex].endswith(SETSTR): 
                jsUp[topics[msgIdIndex]]={}
                jsUp[topics[msgIdIndex]]['bindSigNames'] = bindSigNames
            else:
                jsDown[addSet(topics[msgIdIndex])] = {}
                for bindSigName in bindSigNames:
                    jsDown[addSet(topics[msgIdIndex])][bindSigName]={}
            sigs.clear()

    if len(topics)!=0:    
        writeJs(down_config,jsDown)
        writeJs(up_config,jsUp)     
    else:
        addConfigSig(sigs,True,"",down_config,up_config)

'''
需要信号值转换的就填,格式如下："APP":SIC
"1":2
"2":4
"3":7
defaultValueType:0;  默认值(没有在"值映射"中的值)的处理方式,0是原值转发, 1是不处理,2是加自增1,3是发送无效,4 extension 相同就不发送,5整帧CAN报文透传
'''

def clearXlsData(jsConfig):
    assert isinstance(jsConfig,dict)
    tempJs = copy.deepcopy(jsConfig)
    for fContent in tempJs:
        if type(tempJs[fContent]) == dict:
            for tContent in tempJs[fContent]:
                contentJson = tempJs[fContent][tContent]
                if type(contentJson) == dict and ISXLS in contentJson and contentJson[ISXLS]:
                    del jsConfig[fContent][tContent]
                    if len(jsConfig[fContent]) == 0:
                        del jsConfig[fContent]
                    elif len(jsConfig[fContent]) == 1:
                        if  LOGSTR in jsConfig[fContent]:
                            del jsConfig[fContent]

def addConfigByTopicCanXls(xlsPath,configPath):
    book = xlrd.open_workbook(xlsPath)
    assert isinstance(book, Book)
    sheel = book.sheet_by_index(0)
    addConfigTopicCan(sheel,sheel.nrows,getCellValue,configPath)

def getErrList(errorStrList):
    errorList=[]
    errorList.extend(errorStrList)
    home_dir = os.path.expanduser("~").replace('/home','')
    home_dir = home_dir.replace('/','')
    errorList.append(home_dir)
    current_date = datetime.datetime.now()
    three_days_later = current_date + datetime.timedelta(days=3)
    date_string = three_days_later.strftime("%Y.%m.%d")
    errorList.append(date_string)
    return errorList

def getThreeSig(sheel,rowCount,getSheelValue,rowRange=[],configPath=""): 
    jsDown, jsUp, dbc, jsConfig,down_config,up_config = getJsConfig(configPath)
    for i in range(rowCount):
        if not (i  in rowRange or len(rowRange) == 0):
            continue
        try:
            sigName = getSheelValue(sheel,i,'C').replace(" ", "").replace("\n", "").replace("\r", "")
            if len(sigName) == 0: raise Exception(f"是空的")
        except:
                continue
        configXls = ConfigXls()
        configXls.dbcSigName,configXls.sender,isChanged,messageId = configConverdbc(sigName,dbc,False)
        if configXls.sender == None:
            continue           
        try:
            threeSig = getSheelValue(sheel,i,'I')
            if len(threeSig) != 0:
                print(f'{i} {configXls.dbcSigName:<15} {threeSig}')
        except:
            pass

def getValueRange(valueMap,row):
    assert isinstance(valueMap,str)
    valueMapStrs = valueMap.splitlines()
    valueMapValue=[]
    for valueMapStr in valueMapStrs:
        if valueMapStr.startswith('(') or valueMapStr.startswith('['):
            valueMapStrs = re.findall(e_i,valueMapStr,re.A)
            try:
                minValue = eval(valueMapStrs[0])
                maxValue = eval(valueMapStrs[1]) + 1
                if valueMapStr.startswith('('):
                    minValue = minValue - 1
                if ')' in valueMapStr:
                    maxValue = maxValue - 1
                valueStrs = []
                trValue = valueMapStrs[2]
                print(minValue,maxValue)
                for valueStr in range(minValue,maxValue):
                    valueStrs.append(f'{valueStr}:{trValue}')
                valueMapValue.append("\n".join(valueStrs))
                valueMap = valueMap.replace(valueMapStr,'')
            except :
                printYellow(f'{row+1} 值范围错误')
    if len(valueMapValue) != 0:
        return "\n".join(valueMapValue) + valueMap
    return valueMap

def addConfigTopicCan(sheel,rowCount,getSheelValue,rowRange=[],isAll=False,configPath=""): 
    jsDown, jsUp, dbc, jsConfig,down_config,up_config = getJsConfig(configPath)
    preComments = ""
    whitelistdbcSigNames = []
    book = openpyxl.Workbook()
    sh = book.active
    sh.title = "sheel"
    sh['A1'] = '功能名称'
    sh['B1'] = 'Topic'
    sh['C1'] = '信号名'
    isAll = isAll or len(rowRange) == rowCount or len(rowRange) == 0 
    if isAll:
        clearXlsData(jsDown)
        clearXlsData(jsUp)
    #     writeJs(down_config,jsDown)
    #     writeJs(up_config,jsUp)
    # exit()
    errBook = openpyxl.Workbook()
    errsh = errBook.active
    errsh.title = "sheel"
    errsh['A1'] = '功能'
    errsh['B1'] = '板块'
    errsh['C1'] = '问题来源'
    errsh['D1'] = '问题'
    errsh['E1'] = '提出人'
    errsh['F1'] = '期望解决时间'


    for i in range(rowCount):
        sh.append([])
        if not (i  in rowRange or len(rowRange) == 0):
            continue
        if 1:
        # try:
            noChangedTopic = False
            try:
                customField = int(getSheelValue(sheel,i,xls_ignore))
                if customField == 1:
                    printGreen(f'{i+1:<10}无须导入,有自定义的fds')
                    continue
                noChangedTopic = (customField == 2)
            except:
                pass
            
           
            try:
                #strip()方法用于删除字符串左右两边的空格、特殊字符
                #如果没有指定字符，则默认删除空格以及制表符、回车符、换行符等特殊字符
                sigName = getSheelValue(sheel,i,xls_sig_name).strip()
                if len(sigName) == 0: raise Exception(f"是空的")
            except:
                    printYellow(f'{i+1:<10}行是空的')
                    continue
            configXls = ConfigXls()
            configXls.sigName = sigName
            try:
                value_type = getSheelValue(sheel,i,xls_value_type).strip().upper()
                if value_type == "STRING" or value_type == "MAP": configXls.sigNameType = 1
            except:
                pass

            try:
                configXls.log = int(getSheelValue(sheel,i,xls_log).strip())
            except:
                pass

            configXls.topic = getSheelValue(sheel,i,xls_topic_name).strip()
            if configXls.getBindSigNames(sigName,dbc,whitelistdbcSigNames,jsUp):
                printGreen(f"{i+1:<10} 写入完成")
                continue
            configXls.dbcSigName,configXls.sender,isChanged,messageId = configConverdbc(sigName,dbc)
            configXls.comments = getSheelValue(sheel,i,xls_comments)
            if configXls.comments == "" :
                configXls.comments = preComments
            else:
                preComments = configXls.comments

            if configXls.sender == None:
                errorStrList=[]
                errorStrList.append(configXls.comments)
                try:
                    errorStrList.append(getSheelValue(sheel,i,xls_functional_module))
                except:
                    errorStrList.append('')
                try:
                    errorStrList.append(getSheelValue(sheel,i,xls_prd_name))
                except:
                    errorStrList.append('')
                errorStrList.append(f'{sigName} 信号缺失')
                errsh.append(getErrList(errorStrList))
                continue
           
            # try:
            #     threeSig = getSheelValue(sheel,i,'I')
            #     if len(threeSig) != 0:
            #         print(f'{i} {sigName:<15} {threeSig}')
            # except:
            #     pass
            whitelistdbcSigNames.append(configXls.dbcSigName.replace('/','__'))
            try:
                valueMapStr = getSheelValue(sheel,i,xls_value_mapstr)
                valueMapStr = getValueRange(valueMapStr,i)
                valueMapStrs = re.findall(e_i,valueMapStr,re.A)
                for valueMapStrIndex in range(len(valueMapStrs)):
                    if valueMapStrIndex % 2 != 0 : continue
                    fValue = valueMapStrs[valueMapStrIndex]
                    tIndex = valueMapStrIndex + 1
                    if "defaultValueType" in fValue: continue 
                    if tIndex >= len(valueMapStrs):
                        printYellow(f"{i+1:<10} {tIndex} {len(valueMapStrs)}值映射错误")
                        continue
                    tValue = eval(valueMapStrs[tIndex])
                    configXls.valueMap[fValue] = tValue

                defaultValueTypeStr = re.findall(f'defaultValueType:{i_i}',valueMapStr,re.A)[0]
                assert isinstance(defaultValueTypeStr,str)
                defaultValueTypeStrRow = defaultValueTypeStr.split(":")
                configXls.defaultValueType = defaultValueTypeStrRow[1]
            except:
                pass

            if configXls.sender not in local_machine_Sender: 
                configXls.addToConfig(jsUp,False,noChangedTopic)         
            if configXls.sender in local_machine_Sender:
                configXls.addToConfig(jsDown,True,noChangedTopic) 

            newSheel = configXls.getList()
            sh.cell(row=i, column=1).value = newSheel[0]
            sh.cell(row=i, column=2).value = newSheel[1]
            sh.cell(row=i, column=3).value = newSheel[2]
            printGreen(f"{i+1:<10} 写入完成")

        # except:
        #     printRed(f"{sigName} 信号格式错误")
        #     continue 

    writeJs(down_config,jsDown)
    writeJs(up_config,jsUp)

    book.save("./newTopicCan.xls")
    errBookFile = './errSig.xls'
    if os.path.exists(errBookFile):
        os.remove(errBookFile)
    errBook.save(errBookFile)
    return whitelistdbcSigNames


def printKey(jsConfig,key):
    print(f'{key}:{json.dumps(jsConfig[key],indent=4 ,ensure_ascii=False)}')

def findJson(jsConfig,key):
    assert isinstance(jsConfig,dict)
    for fContent in jsConfig:
        if key in fContent:
            printKey(jsConfig,fContent)
        if type(jsConfig[fContent]) == str:
            if key in jsConfig[fContent]:
                printKey(jsConfig,fContent)
        elif type(jsConfig[fContent]) == dict:
            for tContent in jsConfig[fContent]:
                if key in tContent:
                    printKey(jsConfig,fContent)
        
def findConfigInfo(key,configPath):
    jsDown, jsUp, dbc, jsConfig,down_config,up_config = getJsConfig(configPath)
    printGreen('*************************下行*************************')
    findJson(jsDown,key)
    printGreen('*************************上行*************************')
    findJson(jsUp,key)

if __name__ == "__main__":
    parse = argparse.ArgumentParser(
        description='''
        检查配置文件是否正确
        从配置文件生成java代码
        从java生成配置文件
        ''')
    parse.add_argument('-j', '--ToJave', help='生成json代码', nargs='?', type=str)
    parse.add_argument('-c', '--ToConfig', help='从java代码生成配置文件',type=str)
    parse.add_argument('-s', '--addSig', help='把信号添加带配置文件中',nargs='+',type=str)
    parse.add_argument('-a', '--xls', help='把表格中所有的信号都添加配置文件中,-s为表格的路径',type=str,default='')
    parse.add_argument('-d', '--downjson', help='下行文件路径，没有从配置中读取',type=str,default='', nargs='?')
    parse.add_argument('-u', '--upjson', help='上行文件路径，没有从配置中读取',type=str,default='', nargs='?')
    parse.add_argument('-t', '--topic', help='topic是列表和-m参数索引要对应',default=[], nargs='+', type=str)
    parse.add_argument('-m', '--messages', help='新增messages是一个列表',default=[], nargs='+', type=str)
    parse.add_argument('-tc', '--topicCan', help='通过topic和CAN对应的xls添加配置文件',type=str,default='')
    parse.add_argument('-f', '--findKey', help='发现对应的信息',type=str,default='')
    arg = parse.parse_args()
    configPath = pyFileDir+"config.json"

    if '-m' in sys.argv:
        addMultipleSig(arg.xls,arg.messages,arg.topic)
    elif '-j' in sys.argv:
        ToJaveCode(arg.ToJave,configPath)
    elif '-c' in sys.argv:
        ToConfigJson(arg.ToConfig,configPath)
    elif '-s' in sys.argv:
        addConfigSig(arg.addSig,False,configPath)
    elif '-a' in sys.argv:
        addConfigByCanXls(arg.xls)
    elif '-tc' in sys.argv:
        addConfigByTopicCanXls(arg.topicCan,configPath)
    elif '-f' in sys.argv:
        findConfigInfo(arg.findKey,configPath)
    else:
        CheckSigName(configPath,arg.downjson,arg.upjson)
    

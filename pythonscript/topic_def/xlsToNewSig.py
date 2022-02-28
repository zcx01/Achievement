#!/usr/bin/python
import sys
import os


import xlrd
import time
pyFileDir = os.path.dirname(os.path.abspath(__file__))+"/AnalyzeCan/"
from commonfun import *
import openpyxl 
import argparse
import pyperclip
from AnalyzeCan.Analyzedbc import *

SETSTR='/Set'
jsConfig = getJScontent(pyFileDir+"config.json",)
dbc=Analyze(getKeyPath("dbcfile",jsConfig))

#信号是否存在
#sheel:所有值得集合 - 如一行一行组成的列表
#row:行号
#endCol:最后的列数
#CallFun:得到值得函数
#isUp:判断是否上行
#返回 主信号名、是否存在，关联的信号
def sigExist(sheel,row,endCol,CallFun,isUp):
    sigs=[]
    if type(endCol) == str:
        end = ord(endCol) - ord('A')
    for singCol in range(end):
        sigs.append(str(getValue(CallFun,sheel,row,singCol)))
    relation=''
    exist = False
    for sig in sigs:
        if ',' in sig:
            sigs = sig.split(',')
            sig = sigs[0]
            del sigs[0]
            relation = ','.join(sigs)
        if dbc.sigExist(sig):
            if isUp:
                exist = dbc.isLocateSend(sig)
            else:
                exist = not dbc.isLocateSend(sig)
            # exist = len(sig)>3 and sig!='None'
            # if not exist:
            #     print(row+1,chr(col+65),'信号不存在')
            return sig,exist,relation
    return '',exist,relation

#返回 define topic desc 
#deContent:文件的内容
#contents:匹配的字符串列表 -- 如果是desc就是一个字符串
#spaceIndex:匹配空格的索引，
#如 1匹配字符串被空格分割的第1一个字符,0匹配中文描述
def analyDefine(deContent,contents,spaceIndex):
    desc=''
    defines=[]
    topics=[]
    descs=[]
    for lineContent in deContent:
        if not lineContent.startswith("#define"):
            desc=lineContent
            continue
        is_in = False
        if spaceIndex == 0:
            is_in = desc in contents
        else:
            spaceStr = splitSpaceGetValueByIndex(lineContent,spaceIndex)
            spaceStr = spaceStr.replace('\"','')
            is_in = spaceStr in  contents

        if is_in:
            define=splitSpaceGetValueByIndex(lineContent, 1)
            topic = splitSpaceGetValueByIndex(lineContent, 2)
            if len(contents) == 0:
                return define , topic , desc
            else:
                defines.append(define)
                topics.append(topic)
                descs.append(desc)

    if len(defines) == 0:
        printYellow(f'没有找到 {str(contents)},请更新topic文件')
    return defines,topics,descs

#得到define通过文件define.h
def getDefineByFile(deContent,topic):
    topiclist=[]
    topiclist.append(topic)
    define,topic,dec = analyDefine(deContent,topiclist,2)
    return define[0]

#通过字符串拼接成define
def getDefineBySelf(topic):
    topic = topic.replace('/','')
    define = ''
    for t in topic:
        t = str(t)
        if t.isupper():
            define+='_'+t
        else:
            define+=t.upper()
    return 'IPC'+ define

#生成参数通过表格
def generateByXls(xlsPath,startRow,endRow,down,up):
    book=xlrd.open_workbook(xlsPath)
    sheel=book.sheet_by_index(0)
    generate(sheel,startRow,endRow,down,up,xlsVaule,sheel.nrows)

def xlsVaule(sheel,row,col):
    return sheel.cell_value(row,col)

#得到值
def getValue(CallFun,sheel,row,col):
    try:
        if type(col) == str:
            col = ord(col) - ord('A')
        return CallFun(sheel,row,col)
    except:
        # print(row+1,chr(col+65),'值不存在')
        return

#得到topic通过表格 3行拼接成topic
def getTopicByXls(CallFun,sheel,row):
    topics = []
    try:
        col=0
        findRow = row
        COLNUM=3
        while col < COLNUM:
            topic = CallFun(sheel,findRow,col)
            assert isinstance(topic,str)
            if col == COLNUM-1 and len(topic) == 0:
                break
            if len(topic) != 0:
                #去除第一个topic中的/
                topic = topic.replace('/','')
                topics.append(topic)
                findRow = row
                col+=1
            else:
                findRow-=1
    except:
        if len(topics):
            print(row+1,'值不存在')
    topic = '/'.join(topics)
    return topic+SETSTR,topic

#在第5行获取备注(中文名称) - 状态
def getComments(CallFun,sheel,startRow):
    comment = getValue(CallFun,sheel,startRow,5)
    statusStr='状态'
    commentStatus=comment
    if statusStr not in comment:
        commentStatus=comment+statusStr
    return comment,commentStatus

#得到topic和deine通过行号 来源:表格
#sheel:所有值得集合 - 如一行一行组成的列表
#CallFun:得到值得函数
#desc:中文描述
#rows:指定行号裂变
def getTopicAndDefineByRow(sheel,CallFun,rows):
    subTopic=[]
    descs=[]
    for row in rows:
        row-=1
        commentSet, comment = getComments(CallFun, sheel, row)
        topicStrSet, topicStr = getTopicByXls(CallFun, sheel, row)
        defineStrSet = getDefineBySelf(topicStrSet)
        defineStr = getDefineBySelf(topicStr)
        subTopic.append(topicStr)
        subTopic.append(topicStrSet)
        descs.append(comment)
    interactiveTopic(subTopic,descs)

def interactiveTopic(subTopic,descs,isTip=True,filePaths=[]):
    try:
        for row in range(len(subTopic)):
            print(f'{row:<5}{subTopic[row]:<50}{descs[row]}')
        if len(filePaths) != 0:
            print(filePaths)
        rowIndex = -1
        while True:
            if isTip:
                if len(subTopic) == 1:
                    rowIndex  =0
                else:
                    m = input('是否生成消息(y/n/h)')
                    if  m == 'n':return
                    rowIndex = input('输入索引')
            for row in range(len(subTopic)):
                if row == int(rowIndex) or rowIndex == -1:
                    singleTopic = str(subTopic[row]).replace(r'"','')
                    sendTopic=''
                    if singleTopic.endswith(SETSTR):
                        sendTopic = sendMqtt(f'{singleTopic}',1)
                    else:
                        sendTopic = subMqtt(f'\'{singleTopic}\'')
                    printGreen(sendTopic)
                    pyperclip.copy(sendTopic)
            if not isTip:return
            if len(subTopic) == 1:return
    except:
        pass

#得到topic和deine通过desc 来源:表格
#通过topic表格的comment匹配desc找到行号,然后通过行号找topic和deine
#sheel:所有值得集合 - 如一行一行组成的列表
#CallFun:得到值得函数
#desc:中文描述
#rows:所有的行号
def getTopicAndDefineByDesc(sheel,CallFun,desc,rows):
    findRows=[]
    for row in range(rows):
        try:
            commentSet, comment = getComments(CallFun, sheel, row)
            if desc in commentSet or desc in comment:
                findRows.append(row+1)
        except:
            pass
    getTopicAndDefineByRow(sheel,CallFun,findRows)

#在字符串查找define
#defineSets:define的Set()集合
#content:字符串
def getDefineByContent(defineSets,content):
    defines =  re.findall(d_t, content, re.A)
    for define in defines:
        assert isinstance(define,str)
        defineSets.add(define)
    return defineSets

#在文件源码中得到define通过信号
#sigName:信号名
#srcPath:源码路径
def getInSrcDefineBySig(sigName,srcPath):
    defineSets=set()
    classNames = set()
    filePaths = set()
    for (dirpath,dirnames,filenames) in os.walk(srcPath):
        for fileName in filenames:
            if os.path.splitext(fileName)[1] == '.cpp': 
                filePath = dirpath+'/'+fileName
                contents = readFileLines(filePath)
                # if fileName != "vc_discharge_soc.cpp": continue
                for content in contents:
                    className = getRelationClassName(content)
                    if sigName.upper() in content.upper():
                        getDefineByContent(defineSets,content)
                        if len(className) != 0:
                            classNames.add(className)
                        if len(defineSets) != 0:
                            filePaths.add(filePath)
                            return defineSets,filePaths
   
                with open(filePath, "r") as cr:
                    content = cr.read()
                    if sigName.upper() in content.upper():
                        getDefineByContent(defineSets,content)
                        filePaths.add(filePath)
                        className = getSelfClassName(content)
                        classNames.add(className)

    if len(defineSets) == 0 and len(classNames) !=0:
        for (dirpath,dirnames,filenames) in os.walk(srcPath):
            for fileName in filenames:
                if os.path.splitext(fileName)[1] == '.cpp': 
                    filePath = dirpath+'/'+fileName
                    # if filePath != '/home/chengxiongzhu/Works/Repos/changan_c835/src/ic_service/src/signal_process/src/vehctrl_module.cpp': continue
                    contents = readFileLines(filePath)
                    for contentIndex in range(len(contents)):
                        #if contentIndex != 300 : continue
                        tempclassName = getRelationClassName(contents[contentIndex])
                        if tempclassName in classNames:
                            while contentIndex < len(content):
                                getDefineByContent(defineSets,contents[contentIndex])
                                contentIndex+=1
                                classNames = getRelationClassName(contents[contentIndex])
                                if len(classNames) != 0:
                                        filePaths.add(filePath)
                                        return defineSets,filePaths
                        contentIndex+=1
    return defineSets,filePaths

#通过信号名称topic和deine
def getTopicAndDefineBySig(sigName):
    topic_config_path = pyFileDir+"../topic_config.json"
    print(topic_config_path)
    topic_configs ={}
    try:
        topic_configs = getJScontent(topic_config_path)
    except:
        pass
    realTopic = []
    realDesc = []
    filePaths = []
    config_exist = False
    topic_config = topic_configs.get(sigName,None)
    if topic_config !=  None:
        realTopic = topic_config['topic']
        realDesc = topic_config['desc']
        filePaths = topic_config['filePath']
        config_exist = True

    update = 'y'
    if config_exist:
        interactiveTopic(realTopic,realDesc,False,filePaths)
        update=input("是否更新(y/n)")

    if update == "y": 
        jsConfig = getJScontent(pyFileDir+"config.json")
        srcPath = getKeyPath("srcPath",jsConfig)
        topics=[]
        defineSets,filePaths=getInSrcDefineBySig(sigName,srcPath)
        if len(defineSets) ==0:
            print(f'没有找到')
            return
        defineContents = readFileLines(getKeyPath("definefile",jsConfig)) 
        defines,topics,decs = analyDefine(defineContents,defineSets,1)    
        if len(topics) > 1:
            for row in range(len(topics)):
                print(f'{row:<5}{topics[row]:<50}{decs[row]}')
            rows = input("请输入<以特殊符号(如:,)分割的符号,-1为所有>topic的行号:\n")
            rows = re.findall(e_i,rows,re.A)
            for row in rows:
                row = int(row)
                if(row == -1):
                    realTopic = topics
                    realDesc = decs
                    break
                realTopic.append(topics[row])
                realDesc.append(decs[row])
        else:
            realTopic = topics
            realDesc = decs
        topic_config={}
        topic_config['topic'] = realTopic
        topic_config['desc'] = realDesc
        topic_config['filePath'] = list(filePaths)
        topic_configs[sigName] = topic_config

        writeJs(topic_config_path,topic_configs)
        interactiveTopic(realTopic,realDesc,False,filePaths)

#生成参数
def generate(sheel,startRow,endRow,down,up,CallFun,rows):
    if startRow == 0:
        endRow = rows
    else:
        startRow-=1

    if endRow == -1:
        endRow = startRow
    else:
        endRow-=1

    if startRow >= rows or endRow >= rows or startRow > endRow:
        print('输入的行号不合法')
        return
    jsConfig = getJScontent(pyFileDir+"config.json")
    defineContents = readFileLines(getKeyPath("definefile",jsConfig)) 
    xlsNewSigPath = getKeyPath("xlsNewSigPath",jsConfig)
    book = openpyxl.Workbook()
    rowContent=[]
    sh = None

    if os.path.isfile(xlsNewSigPath):
        book = openpyxl.load_workbook(xlsNewSigPath)
        sheetsSize = len(book.sheetnames)
        sh = book[book.sheetnames[0]]

        if sheetsSize > 1:
            old_sh = book[book.sheetnames[sheetsSize-1]]
            old_sh.append([time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())])
            for row in range(sh.max_row):
                if row == 0:
                    continue
                temp = []
                for col in range(sh.max_column):
                    temp.append(sh.cell(row+1,col+1).value)

                old_sh.append(temp)
            sh.delete_rows(0,sh.max_row+1)
    else:
        book = openpyxl.Workbook()
        sh = book.active

    rowContent=[]
    rowContent.append('信号')
    rowContent.append('类型')
    rowContent.append('中文名称(d)')
    rowContent.append('类名(x)')
    rowContent.append('topic(x表示自动搜索)')
    rowContent.append('上行信号/topic')
    rowContent.append('关联信号')
    sh.append(rowContent)

    while(startRow <= endRow):
        commentSet,comment =  getComments(CallFun,sheel,startRow)
        topicStrSet, topicStr = getTopicByXls(CallFun, sheel, startRow)
        className = getValue(CallFun,sheel,startRow,1)+ getValue(CallFun,sheel,startRow,2)
        sig,isExist,relation = sigExist(sheel,startRow,'N',CallFun,True)
        isTopic = True
        if isExist:
            isTopic = False
            rowContent=[]
            topicDefine = getDefineByFile(defineContents,topicStrSet)
            rowContent.append(sig)
            rowContent.append(down)
            rowContent.append(commentSet)
            rowContent.append(className)
            rowContent.append(topicDefine)
            rowContent.append('n')
            rowContent.append(relation)
            print(rowContent)
            sh.append(rowContent)

        sig,isExist,relation = sigExist(sheel,startRow,'N',CallFun,False)
        if isExist:
            isTopic = False
            rowContent=[]
            topicDefine = getDefineByFile(defineContents,topicStr)
            rowContent.append(sig)
            rowContent.append(up)
            rowContent.append(comment)
            rowContent.append(className+'Status')
            rowContent.append(topicDefine)
            rowContent.append('y')
            rowContent.append(relation)
            print(rowContent)
            sh.append(rowContent)

        if isTopic:
            rowContent=[]
            tempSet = str(getValue(CallFun,sheel,startRow,'H'))
            if isNumber(tempSet):
                topicStrSet, tempStr = getTopicByXls(CallFun, sheel, tempSet)
                topicDefineSet = getDefineByFile(defineContents,topicStrSet)
            elif len(tempSet) == 0:
                topicDefineSet = getDefineByFile(defineContents,tempSet)
            else:
                topicDefineSet = tempSet
            topicDefine = getDefineByFile(defineContents,topicStr)
            rowContent.append(topicDefineSet)
            rowContent.append(up)
            rowContent.append(comment)
            rowContent.append(className)
            rowContent.append(topicDefine)
            rowContent.append('-b')
            rowContent.append(relation)
            print(rowContent)
            sh.append(rowContent)
        startRow+=1

    saveFileName=xlsNewSigPath
    book.save(saveFileName)
    book.close()
    print("生成完成")
    os.system(f'xdg-open {saveFileName}')

# getTopicAndDefineBySig('CdcWiprSrvSig')
if __name__ == "__main__":
    parse = argparse.ArgumentParser(description='这个是通过topic表格生成生成newSig表格')
    parse.add_argument('-d','--down',help='下行信号的类型',default="vehctrl")
    parse.add_argument('-u','--up',help='上行信号的类型',default="vehctrl_status")
    parse.add_argument('-x','--xlsPath',help='topic表格路径',default='/home/chengxiongzhu/Works/newsig/topic.xls')
    parse.add_argument('-s','--startRow',help='开始的行号',type=int,default=0)
    parse.add_argument('-e','--endRow',help='结束的行号',type=int,default=-1)
    parse.add_argument('-f','--findName',help='信号和topic互转',type=str)
    arg = parse.parse_args()
    if '-f' in sys.argv:
        getTopicAndDefineBySig(arg.findName)
    else:
        generateByXls(arg.xlsPath,arg.startRow,arg.endRow,arg.down,arg.up)

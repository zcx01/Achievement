#!/bin/python3
import os
import sys
import argparse
from commonfun import *
import openpyxl 
from xml.dom.minidom import *
import xlrd
from xlrd.book import Book
from xlrd.sheet import Sheet

SOUCRELANGUAGESUFFIX="zh_CN.ts"
ORILANGUAGESUFFIXS={
                "en_US.ts":"C",
                "th_TH.ts":"D",
                "ar_EG.ts":"E",
                "da_DK.ts":"F",
                "de_DE.ts":"G",
                "ru_RU.ts":"H",
                "fr_FR.ts":"I",
                "fi_FI.ts":"J",
                "nl_NL.ts":"K",
                "pt_PT.ts":"L",
                "sv_SE.ts":"M",
                "es_ES.ts":"N",
                "it_IT.ts":"O",
                "nb_NO.ts":"P",
                    }
 
ORILANGUAGENAME={
                "en_US.ts":"英文",
                "th_TH.ts":"泰文",
                "ar_EG.ts":"阿拉伯语",
                "da_DK.ts":"丹麦语",
                "de_DE.ts":"德语",
                "ru_RU.ts":"俄语",
                "fr_FR.ts":"法语",
                "fi_FI.ts":"芬兰语",
                "nl_NL.ts":"荷兰语",
                "pt_PT.ts":"葡萄牙语",
                "sv_SE.ts":"瑞典语",
                "es_ES.ts":"西班牙语",
                "it_IT.ts":"意大利语",
                "nb_NO.ts":"挪威语",
                } 
SHELLNAME="SIC"

def getExistranslateCols():
    if len(ORILANGUAGESUFFIXS) == 0:
        return "C"
    exisTranslates = list(ORILANGUAGESUFFIXS.values())
    exisTranslates.sort()
    return exisTranslates[-1]

EXISTRANSLATEMAXCOLS = XlsCharToInt(getExistranslateCols()) #翻译最大的行数
EXISTRANSLATESTARTCOL = 2   #翻译开始的行数
class translateConent(object):
    def __init__(self) -> None:
        self.sourceKey=""
        self.content=""
        self.exisTranslatesList=[]
        self.translateContent=""
        self.changedCount = EXISTRANSLATESTARTCOL
        for i in range(EXISTRANSLATESTARTCOL,EXISTRANSLATEMAXCOLS+1):
            self.exisTranslatesList.append('')

    def getList(self):
        xlsContent = []
        xlsContent.append(self.sourceKey)
        xlsContent.append(self.content)
        for exisTranslate in self.exisTranslatesList:
            xlsContent.append(exisTranslate)
        return xlsContent
    
    def appendExis(self,key,exisTranslate):
        if len(exisTranslate) != 0:
            self.exisTranslatesList[XlsCharToInt(ORILANGUAGESUFFIXS[key])-EXISTRANSLATESTARTCOL] = exisTranslate
            self.changedCount = self.changedCount+1

    def appendSh(self,sh,isCheckChange):
        if not isCheckChange:
            sh.append(self.getList())
        elif self.changedCount < EXISTRANSLATEMAXCOLS and len(self.content) != 0:
            sh.append(self.getList())
            
    
def getTextCout(fileName):
    jsContents = getJScontent(fileName)
    count=0
    for top in jsContents:
        for grade in jsContents[top]:
            if isNumber(grade) :
                print(EesyStr.removeAll(jsContents[top][grade],'\n'))
                count+=1

def joint(f,t):
    return f'{f}__{t}'

def getValue(src, row, col):
    return src.cell_value(row,XlsCharToInt(col))

def removeSuffix(conetnt):
    return conetnt.replace(".ts","")

def translatesByTs(dirpath,oriName,translateConents,repeatTranslateConents,isSource):
    translatePath_file = f'{dirpath}/{oriName}'
    if not os.path.exists(translatePath_file): return False
    domTree = parse(translatePath_file)
    assert isinstance(domTree,Document)
    rootNode = domTree.documentElement
    assert isinstance(rootNode,Element)
    contexts = rootNode.getElementsByTagName("context")
    for context in contexts:
        assert isinstance(context,Element)
        contentFile = context.getElementsByTagName("name")[0].childNodes[0].data
        messageNodes = context.getElementsByTagName("message")
        for messageNode in messageNodes:
            assert isinstance(messageNode,Element)
            try:
                source = messageNode.getElementsByTagName("source")[0].childNodes[0].data
            except:
                continue
            try:
                content= messageNode.getElementsByTagName("translation")[0].childNodes[0].data
            except:
                if isSource:
                    content = source
                else:
                    content=""

            xlsContent = translateConent()
            xlsContent.sourceKey = source
            xlsContent.content = content

            if repeatTranslateConents == None:
                translateConents[source] = xlsContent
            elif source not in repeatTranslateConents:
                translateConents[source] = xlsContent
                repeatTranslateConents[source] = xlsContent
    return True            

def coverXls(translatePath,jsonPath,outPath,isCheckChange=False):
    # 创建一个Excel workbook 对象
    # if os.path.isfile(outPath):
    #     book = openpyxl.load_workbook(outPath)
    #     sh = book['English']
    # else:
    book = openpyxl.Workbook()
    sh = book.active
    sh.title = SHELLNAME
    sh['A1'] = '源key'
    sh['B1'] = '内容'
    for exisTranslate,col in ORILANGUAGESUFFIXS.items():
        sh[f'{col}1'] = ORILANGUAGENAME[exisTranslate]

    exisTranslates = ORILANGUAGESUFFIXS.keys()

    repeatSoucreLanguages={}
    for (dirpath,dirnames,filenames) in os.walk(translatePath):
        for oriName in filenames:
            if SOUCRELANGUAGESUFFIX in oriName:
                soucreLanguages={}
                translatesByTs(dirpath,oriName,soucreLanguages,repeatSoucreLanguages,True)
                existranslatelanguages = {}

                changedExisTranslates = [] #去除不存在的语言
                for exisTranslate in exisTranslates:
                    existransName = oriName.replace(SOUCRELANGUAGESUFFIX,exisTranslate)
                    existranslatelanguage = {}
                    if translatesByTs(dirpath,existransName,existranslatelanguage,None,False):
                        existranslatelanguages[exisTranslate] = existranslatelanguage
                        changedExisTranslates.append(exisTranslate)
                exisTranslates = changedExisTranslates
                count = 0
                for soucreLanguage in soucreLanguages:
                    xlsContent=soucreLanguages[soucreLanguage]
                    assert isinstance(xlsContent,translateConent)
                    for exisTranslate in exisTranslates:
                        try:
                            xlsContent.appendExis(exisTranslate,existranslatelanguages[exisTranslate][xlsContent.sourceKey].content)
                        except:
                            # printYellow(f"{xlsContent.sourceKey} 不存在 {removeSuffix(exisTranslate)}")
                            pass
                    xlsContent.appendSh(sh,isCheckChange)
                    count+=1
                print(f"检测完 {dirpath}/{oriName} 文件 一共有 {count}")
    
    try:
        exisTranslates = ORILANGUAGESUFFIXS.keys()
        jsContents = getJScontent(jsonPath)
        count=0
        for top in jsContents:
            for grade in jsContents[top]:
                if isNumber(grade) :
                    xlsContent = translateConent()
                    xlsContent.sourceKey = joint(top,grade)          
                    xlsContent.content = jsContents[top][grade][removeSuffix(SOUCRELANGUAGESUFFIX)]
                    for exisTranslate in exisTranslates:
                        try:
                            xlsContent.appendExis(exisTranslate,jsContents[top][grade][removeSuffix(exisTranslate)])
                        except:
                            printYellow(f"{top} {grade}不存在 {removeSuffix(exisTranslate)}")
                    xlsContent.appendSh(sh,isCheckChange)
                    count+=1
        print(f"检测完 {jsonPath} 文件 一共有{count}")
    except:
        printRed(f"{jsonPath} 异常")
        pass
    book.save(outPath)
    printGreen("检测完成")

#只有修改xls转化成输入的xls
def getChangeTrInputXls(changeTrInput):
    book = xlrd.open_workbook(changeTrInput)
    assert isinstance(book, Book)

    sheel = book.sheet_by_name(SHELLNAME)
    assert isinstance(sheel, Sheet)

    book = openpyxl.Workbook()
    sh = book.active

    for row in range(sheel.nrows):
        moudleName = getValue(sheel,row,"A")
        if moudleName == 'IC-OUT' or moudleName == "SIC" or moudleName == ".json" or moudleName == '.ts' or row == 1:
            sh.append(row)
    return sh

    
def xlsCover(translatePath,jsonPath,input):
    book = xlrd.open_workbook(input)
    assert isinstance(book, Book)

    sheel = book.sheet_by_name(SHELLNAME)
    assert isinstance(sheel, Sheet)

    for oriLanguageSuffix,oriLanguageSuffixCol in ORILANGUAGESUFFIXS.items():
        tsContent={}
        printGreen(f"正在更新 {oriLanguageSuffix} ...")
        for row in range(sheel.nrows):
            xlsContent = translateConent()
            xlsContent.sourceKey = getValue(sheel,row,"A")
            xlsContent.content = getValue(sheel,row,"B")
            # xlsContent.translateContent = getValue(sheel,row,"C")
            assert isinstance(xlsContent.translateContent,str)
            try:
                newTranslateContent = str(getValue(sheel,row,oriLanguageSuffixCol))
            except:
                if row == 0:
                    break
                else:
                    continue
            if len(xlsContent.sourceKey) == 0:
                xlsContent.sourceKey = xlsContent.content
            if len(newTranslateContent)!=0:
                xlsContent.translateContent = newTranslateContent
            if len(xlsContent.translateContent.split('\n')) > 1:
                pass
            else:
                xlsContent.translateContent = xlsContent.translateContent.replace(',',',\n')

            tsContent[xlsContent.sourceKey] = xlsContent
        if len(tsContent) == 0:
            break
        
        #创建文件
        for (dirpath,dirnames,filenames) in os.walk(translatePath):
            for oriName in filenames:
                if SOUCRELANGUAGESUFFIX in oriName:
                    translatePath_file = f'{dirpath}/{oriName}'
                    ori_translatePath_file = translatePath_file.replace(SOUCRELANGUAGESUFFIX,oriLanguageSuffix)
                    if not os.path.isfile(ori_translatePath_file):
                        os.system(f"cp {translatePath_file} {ori_translatePath_file}")
        
        for (dirpath,dirnames,filenames) in os.walk(translatePath):
            for oriName in filenames:
                if oriLanguageSuffix in oriName:
                    print(f"更新 {translatePath_file} 文件")
                    translatePath_file = f'{dirpath}/{oriName}'
                    domTree = parse(translatePath_file)
                    assert isinstance(domTree,Document)
                    rootNode = domTree.documentElement
                    assert isinstance(rootNode,Element)
                    contexts = rootNode.getElementsByTagName("context")
                    for context in contexts:
                        assert isinstance(context,Element)
                        contentFile = context.getElementsByTagName("name")[0].childNodes[0].data
                        messageNodes = context.getElementsByTagName("message")
                        for messageNode in messageNodes:
                            assert isinstance(messageNode,Element)
                            try:
                                source = messageNode.getElementsByTagName("source")[0].childNodes[0].data
                            except:
                                continue
                            if source in tsContent:
                                if len(messageNode.getElementsByTagName("translation")[0].childNodes) == 0:
                                    phone_text_value = domTree.createTextNode(tsContent[source].translateContent)
                                    messageNode.getElementsByTagName("translation")[0].appendChild(phone_text_value)
                                else:
                                    try:
                                        messageNode.getElementsByTagName("translation")[0].childNodes[0].data = tsContent[source].translateContent
                                    except:
                                        continue
                    with open(translatePath_file, 'w') as f:
                        # 缩进 - 换行 - 编码
                        domTree.writexml(f,encoding='utf-8')

        try:
            aimJsContents = getJScontent(jsonPath)  
            print(f"更新 {jsonPath} 文件")
            for top in aimJsContents:
                for grade in aimJsContents[top]:
                    if isNumber(grade) :
                        source = joint(top,grade)  
                        if source in tsContent:
                            aimJsContents[top][grade][removeSuffix(oriLanguageSuffix)] = tsContent[source].translateContent
            writeJs(jsonPath,aimJsContents)
        except:
            pass
    printGreen("更新完成")

def buildScript():
    for key in ORILANGUAGESUFFIXS:
        ts = 'ic_qt_' + key
        print(r'$${PWD}/../resources/translate/'+ts+'\\')

    print('--------------------------')
    for key in ORILANGUAGESUFFIXS:
        qm = 'ic_qt_' + key.replace('.ts','')
        print(f'/opt/qt/translate/{qm}.qm=../mega/prebuilts/bigsur/ic/qt/translate/{qm}.qm')

#./ic_text.py -t /home/chengxiongzhu/Works/Repos/changan_c385/qt/ic_qt/resources/translate -j /home/chengxiongzhu/Works/Repos/changan_c385/qt/ic_qt/resources/config/icwarning_config.json -o ic-out.xlsx
if __name__ == "__main__":
    aparse = argparse.ArgumentParser(description='python的脚本模板')
    aparse.add_argument('-t','--translatePath',help='翻译文件所在目录',type=str,nargs="?",default="/home/chengxiongzhu/Works/Repos/changan_c385/qt/ic_qt/resources/translate")
    aparse.add_argument('-j','--json',help='报警文字配置文件',type=str,nargs="?",default="/home/chengxiongzhu/Works/Repos/changan_c385/qt/ic_qt/resources/config/icwarning_config.json")
    aparse.add_argument('-o','--outPath',help='生成xls文件路径',type=str,nargs="?",default='./changed.xlsx')
    aparse.add_argument('-i','--input',help='输入xls文件路径',type=str)
    aparse.add_argument('-c','--checkChange',help='检测修改的文本',type=str,nargs="?")
    aparse.add_argument('-b','--buildScript',help='生成部署命令',nargs="?")
    arg = aparse.parse_args()

    if len(sys.argv) == 1:
        getTextCout(arg.fileName)
    elif "-o" in sys.argv:
        coverXls(arg.translatePath,arg.json,arg.outPath)
    elif "-i" in sys.argv:
        xlsCover(arg.translatePath,arg.json,arg.input)
    elif '-c' in sys.argv:
        coverXls(arg.translatePath,arg.json,arg.checkChange,True)
    elif "-b" in sys.argv:
        buildScript()
